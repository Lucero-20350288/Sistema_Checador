import platform
import subprocess
from flask import Flask, render_template, request, Response, jsonify, redirect, url_for,make_response,flash
import database as dbase
import pymongo
from flask import send_file
from department import Department
from empleado import Empleado
from datetime import datetime,timedelta
import pandas as pd
from pandas import DataFrame
from openpyxl import Workbook
from io import BytesIO
import pdfkit
import pytz
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment,PatternFill
from bson.json_util import dumps
from incidencia import Incidencia
from bson import ObjectId, Binary
import base64
from bson.regex import Regex
from bson.objectid import ObjectId
from datetime import datetime,timedelta,timezone,date
import re
from fuzzywuzzy import fuzz
from job_position import JobPosition
import os

db = dbase.dbConnection()

app = Flask(__name__)
app.secret_key = os.urandom(24)
#RUTA PRINCIPAL- NO BORRAR
@app.route('/')
def principal():
  empleados = db['datos_generales']
  empleadosReceived = empleados.find()
  return render_template(['index.html'], empleados = empleadosReceived)

#Rutas de la aplicacion

####################################### Routes Justificaciones (Rocio)  ##########################################
@app.route('/incidencias')
def incidencias():
  return render_template('incidencias.html')

#PoUp Se Agrego Correctamente
@app.route('/incidencia_exitosa')
def incidencia_exitosa():
  return render_template('incidencia_exitosa.html')

@app.route('/frameIncidencias')
def frameIncidencias():
    incidencias = db['incidencias']
    incidenciasReceived = incidencias.find()
    return render_template('incidencias.html', incidencias=incidenciasReceived)

@app.route('/Incidencia', methods={'POST'})
def addIncidencia():
  incidencias_collection = db['incidencias']

  # Check if the 'incidencia' key is in the form data
  if 'incidencia' in request.form:
      descripcion = request.form['incidencia']
      
      # Check if descripcion is not empty
      if descripcion:
          # Create an Incidencia object and insert it into the collection
          incidencias = Incidencia(descripcion=descripcion)
          incidencias_collection.insert_one(incidencias.toDB())
          
          # Convert the Incidencia object to a dictionary
          incidencia_dict = incidencias.toDB()
          
          # Return a response or redirect to another route
          response = jsonify({
              'incidencia': incidencia_dict
          })
          return redirect(url_for('incidencia_exitosa'))
      else:
          return 'Error: descripcion cannot be empty', 400
  else:
      return 'Error: Missing incidencia field', 400
    
@app.route('/delete/<string:incidencia_incidencia>')
def delete_inc(incidencia_incidencia):
    department = db['incidencias']
    department.delete_one({'incidencia': incidencia_incidencia})
    return redirect(url_for('frameIncidencia'))
  
  
@app.route('/obtener-incidencias', methods=['GET'])
def obtener_incidencias():
    incidencias_collection = db['incidencias']
    incidencias = list(incidencias_collection.find())  # Convertir a lista para evitar problemas con la iteración

    # Convertir el ObjectId a cadena en cada incidencia
    for incidencia in incidencias:
        incidencia['_id'] = str(incidencia['_id'])

    return jsonify(incidencias)


def incidencia_existe(incidencia):
    # Buscar la incidencia en la colección
    incidencias_collection = db['incidencias']
    incidencias = incidencias_collection.find({}, {'incidencia': 1})

    # Calcular la similitud entre la incidencia ingresada y las existentes
    for incidencia_existente in incidencias:
        similitud = fuzz.ratio(incidencia.lower(), incidencia_existente['incidencia'].lower())
        if similitud > 85:  # Umbral de similitud
            return True  # La incidencia es similar a una existente

    return False  # La incidencia no es similar a ninguna existente
  
@app.route('/verificar-incidencia', methods=['POST'])
def verificar_incidencia():
    incidencia = request.form.get('incidencia')
    
    # Verificar si la incidencia ya existe en la base de datos (debes implementar esta lógica)
    existe_incidencia = incidencia_existe(incidencia)

    # Devolver una respuesta JSON con el resultado de la verificación
    return jsonify({'existe': existe_incidencia})

#frameAutorizacion
@app.route('/frameAutorizacion')
def frameAutorizacion():
  return render_template('justificaciones.html') 

#API para autocompletado
@app.route('/autocomplete-datos', methods=['GET'])
def autocomplete_datosJustificacion():

  empleados = db['datos_generales']
  cursor = empleados.find({}, {'_id':0})
  #resultado = []
  #for index, empleado in enumerate(cursor):
    #nombre = empleado["nombre"]
    #apellido_pat = empleado["apellido_paterno"]
    #apellido_mat = empleado["apellido_materno"]
  return jsonify(list(cursor))


#Obtener departamento del empleado
@app.route('/obtener-departamento', methods=['GET'])
def obtener_departamento():
    
    rfc = request.args.get('rfc')  # Obtener el RFC de la solicitud GET
    
    departamento_collection = db['departamentos_area']

    # Buscar el departamento en la base de datos por el RFC
    departamentoArea = departamento_collection.find_one({'RFC': rfc}, {'departamento_o_area': 1})
    print(departamentoArea)
    if departamentoArea:
        return jsonify({'departamentoArea': departamentoArea['departamento_o_area']}), 200
    else:
        return jsonify({'message': 'No se encontró el departamento'}), 404


#Asignar justificacion
@app.route('/agregar-justificacion', methods=['POST'])
def agregar_justificacion():
    # Recibe los datos del formulario como JSON
    data = request.get_json()

    # Extrae los datos necesarios para la justificación
    rfc = data.get('rfc')
    nombre = data.get('nombre')
    apellido_paterno = data.get('apellido_paterno')
    apellido_materno = data.get('apellido_materno')
    incidencia = data.get('incidencia')
    tipo_fecha = data.get('tipo_fecha')  # Nuevo campo para indicar el tipo de fecha (rangoFechas o fechaUnica)
    fecha_inicial = data.get('fecha_inicial')
    fecha_final = data.get('fecha_final') if tipo_fecha == 'rangoFechas' else fecha_inicial
    departamento_o_area = data.get('departamento_o_area')

    # Crea el documento de la justificación para insertar en la base de datos
    justificacion_document = {
        'RFC': rfc,
        'nombre': nombre,
        'apellido_paterno': apellido_paterno,
        'apellido_materno': apellido_materno,
        'incidencia': incidencia,
        'fecha_inicial': fecha_inicial,
        'fecha_final': fecha_final,
        'departamento_o_area': departamento_o_area
    }

    try:
        # Inserta el documento en la colección 'justificaciones'
        resultado = db['justificaciones'].insert_one(justificacion_document)

        # Si se inserta con éxito, devuelve el ID del nuevo documento
        return jsonify({'mensaje': 'Justificación agregada exitosamente', 'id': str(resultado.inserted_id)}), 201
    except pymongo.errors.OperationFailure as e:
        # Si hay un error, devuelve un mensaje con el error
        return jsonify({'mensaje': 'Error al agregar justificación', 'error': str(e)}), 500


@app.route('/obtener-justificaciones-por-rfc', methods=['GET'])
def obtener_justificaciones_por_rfc():
    rfc = request.args.get('rfc')  # Obtener el RFC de la solicitud GET

    # Buscar las justificaciones en la base de datos por el RFC
    justificaciones = db['justificaciones'].find({'RFC': rfc})

    # Crear una lista de justificaciones en formato JSON
    justificaciones_json = []
    for justificacion in justificaciones:
        justificacion_json = {
            'incidencia': justificacion['incidencia'],
            'fecha_inicial': justificacion['fecha_inicial'],
            'fecha_final': justificacion['fecha_final']
        }
        justificaciones_json.append(justificacion_json)

    return jsonify(justificaciones_json)  # Devolver las justificaciones como JSON


@app.route('/delete-justificacion', methods=['DELETE'])
def deleteJustificacion():
    data = request.get_json()
    rfc = data.get('rfc')
    incidencia = data.get('incidencia')
    fecha_inicial = data.get('fecha_inicial')

    print(rfc)
    print(incidencia)
    print(fecha_inicial)
    justificaciones_collection = db['justificaciones']
    result = justificaciones_collection.delete_one({'RFC': rfc, 'incidencia': incidencia, 'fecha_inicial': fecha_inicial})

    if result.deleted_count == 1:
        return jsonify({'message': 'Justificación eliminada correctamente'}), 200
    else:
        return jsonify({'message': 'No se encontró la justificación'}), 404



@app.route('/update-justificacion/<rfc>', methods=['PUT'])
def update_justificacion(rfc):
    justificaciones_collection = db['justificaciones']

    # Obtener los datos enviados en la solicitud PUT
    data = request.get_json()

    # Actualizar la justificación usando RFC y posiblemente otros campos para identificación
    result = justificaciones_collection.update_one(
        {'RFC': rfc},  # Asegura la combinación correcta para el filtro
        {'$set': {
            'incidencia': data.get('incidencia'),
            'fecha_inicial': data.get('fecha_inicial'),
            'fecha_final': data.get('fecha_final')
        }}
    )

    if result.modified_count > 0:  # Usar '> 0' para verificar si se realizó algún cambio
        return jsonify({'message': 'Justificación modificada correctamente'}), 200
    else:
        return jsonify({'message': 'No se encontró la justificación o no se realizó ningún cambio'}), 404
    


@app.route('/justificacion-todos')
def justificacion_todos():
    return render_template('justificacion-todos.html')





@app.route('/obtener-empleados')
def obtener_empleados():
    incidencia = request.args.get('incidencia')
    fecha_inicio = request.args.get('fecha_inicial')
    fecha_fin = request.args.get('fecha_final')
    
    print(f"Incidencia: {incidencia}, Fecha inicio: {fecha_inicio}, Fecha fin: {fecha_fin}")


    # Realiza la consulta a tu base de datos para obtener los empleados con los criterios dados
    empleados = db['justificaciones'].find({'incidencia': incidencia, 'fecha_inicial': fecha_inicio, 'fecha_final': fecha_fin})

    # Formatea los resultados como una lista de diccionarios, donde cada diccionario representa un empleado
    empleados_list = []
    for empleado in empleados:
        empleados_list.append({
            'nombre': empleado['nombre'],
            'incidencia': empleado['incidencia'],
            'fecha_inicial': empleado['fecha_inicial'],
            'fecha_final': empleado['fecha_final']
        })

    # Devuelve los empleados como respuesta en formato JSON
    return jsonify(empleados_list)

######################################### Fin de routes rocio #########################################

@app.route('/framedepartment')
def framedepartment():
  return render_template('department.html') 




 ###-------------asd-as-das-d-Asd_asda-sd-sd-das-SDa-Sdas-d AVISOS  CARLOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOOS RUTAS
@app.route('/avisos')
def avisos():
    # Obtener los datos de la colección 'datos_generales'
    datos_generales = list(db.datos_generales.find())

    # Obtener los RFC y nombres de todos los empleados existentes
    rfc_nombres_empleados = [(item['RFC'], item['nombre']) for item in datos_generales]

    return render_template('avisos.html', rfc_nombres_empleados=rfc_nombres_empleados)




@app.route('/frameAvisos', methods=['GET'])
def frameAvisos():
    aviso = request.args.get('aviso')
    rfc = request.args.get('rfc')
    fecha_creacion = request.args.get('fecha_creacion')
    fecha_vencimiento = request.args.get('fecha_vencimiento')
    return render_template('avisos.html', aviso=aviso, rfc=rfc, fecha_creacion=fecha_creacion, fecha_vencimiento=fecha_vencimiento)





@app.route('/frameHistorial')
def frameHistorial():
    return render_template('historial.html')

# Guardar Aviso
@app.route('/guardarAvisos', methods=['POST'])
def guardar_avisos():
    # Obtener los datos enviados en la solicitud POST
    Avisos_data = request.get_json()
    
    # Aquí podrías insertar avisos_data en tu base de datos
    # Por ejemplo:
    db.Avisos.insert_one(Avisos_data)

    # Devolver una respuesta para indicar que fue exitoso
    return jsonify({'mensaje': 'Aviso guardado exitosamente'}), 200

# Guardar AvisoGeneral
@app.route('/guardarAvisoGeneral', methods=['POST'])
def guardar_avisogeneral():
    # Obtener los datos enviados en la solicitud POST
    Avisogeneral_data = request.get_json()
    
    # Aquí podrías insertar avisos_data en tu base de datos
    # Por ejemplo:
    db.Avisogeneral.insert_one(Avisogeneral_data)

    # Devolver una respuesta para indicar que fue exitoso
    return jsonify({'mensaje': 'Aviso guardado exitosamente'}), 200

@app.route('/autocomplete-datos', methods=['GET'])
def autocomplete_datos():

  empleados = db['datos_generales']
  cursor = empleados.find({}, {'_id':0})
  #resultado = []
  #for index, empleado in enumerate(cursor):
    #nombre = empleado["nombre"]
    #apellido_pat = empleado["apellido_paterno"]
    #apellido_mat = empleado["apellido_materno"]
  return jsonify(list(cursor))

    ##LLAMADA PARA BUSCAR EL AVISO

# Ruta para eliminar un aviso
@app.route('/eliminar_aviso/<string:aviso>', methods=['DELETE'])
def eliminar_aviso(aviso):
    db['Avisos'].delete_one({'Aviso': aviso})
    return redirect(url_for('buscaravisos'))


@app.route('/eliminar_aviso_rfc/<string:rfc>', methods=['DELETE'])
def eliminar_aviso_rfc(rfc):
    db['Avisos'].delete_many({'RFC': rfc})  # Elimina todos los avisos con el RFC dado
    return 'Avisos eliminados exitosamente', 200


# Ruta principal para mostrar avisos
@app.route('/buscaravisos', methods=['GET'])
def buscaravisos():
    # Obtener parámetros de búsqueda si los hay
    rfc = request.args.get('rfc')
    # Puedes agregar más parámetros de búsqueda según lo necesites

    # Realizar la consulta a la base de datos de avisos
    avisos_collection = db['Avisos' 'Avisogeneral']
    query = {}  # Consulta vacía para obtener todos los avisos por defecto
    if rfc:
        query['RFC'] = rfc
    # Puedes agregar más condiciones a la consulta según los parámetros recibidos

    cursor = avisos_collection.find(query)

    # Procesar los resultados de la consulta
    resultados = []
    for aviso in cursor:
        resultado = {
            'aviso': aviso['Aviso'],
            'RFC': aviso['RFC'],
            'fecha_creacion': aviso['Fecha de creacion'],
            'fecha_vencimiento': aviso['Fecha de vencimiento']
        }
        resultados.append(resultado)

    # Renderizar la plantilla HTML con los resultados
    return render_template('historial.html', resultados=resultados)

    ##FIN LLAMADA







#--------------sdf-sdf-sd-fs-df-sdf CARLOS AVISOS NO BORRAR -------------------------------------------








#RUTAS DE TOÑO-----------------------------------------------------------------------------------------------EMPIEZAN LAS RUTAS DE TOÑO, NO MODIFICAR-----------------------------------------------------------------


 
@app.route('/frameChecador', methods=['GET', 'POST'])
def frameChecador():
    resultados = []
    if request.method == 'POST':
        rfc = request.form.get('rfc')
        nombre_completo = request.form.get('nombre')
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin') or fecha_inicio

        # Verificar si no se han ingresado datos en los campos de búsqueda
        if not (rfc or nombre_completo or fecha_inicio):
            return render_template('checador.html', resultados=[], error="Ingresa algún dato para realizar la búsqueda.")

        fecha_inicio_dt = None
        fecha_fin_dt = None
        if fecha_inicio and fecha_fin:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)

        if nombre_completo:
            rfc_obtenido = obtener_rfc_por_nombre_completo(nombre_completo)
            if rfc_obtenido:
                rfc = rfc_obtenido  # Sobrescribe el rfc si se encuentra uno mediante el nombre
            else:
                return render_template('checador.html', resultados=[], error="No se encontró un empleado con el nombre proporcionado.")

        query = {"RFC": rfc} if rfc else {}
        registros = db['catalogos_horario'].find(query)

        for registro in registros:
            nombre_completo = obtener_nombre_completo(registro.get('RFC'))
            for fecha in registro.get('Fechas', []):
                fecha_dia_str = fecha.get('fecha_dia', "")
                if fecha_dia_str:
                    fecha_dia_dt = datetime.strptime(fecha_dia_str[:10], '%Y-%m-%d')
                    if fecha_inicio_dt and fecha_fin_dt:
                        if not (fecha_inicio_dt <= fecha_dia_dt <= fecha_fin_dt):
                            continue
                    resultados += procesar_datos_checador(fecha, nombre_completo, registro.get('RFC', 'Desconocido'))

    return render_template('checador.html', resultados=resultados)


def obtener_nombre_completo(rfc):
    """Retrieves the full name from the database based on the RFC."""
    datos_personales = db['datos_generales'].find_one({"RFC": rfc})
    if datos_personales:
        return f"{datos_personales.get('nombre', '')} {datos_personales.get('apellido_paterno', '')} {datos_personales.get('apellido_materno', '')}"
    return "Desconocido"

def obtener_rfc_por_nombre_completo(nombre_completo):
    partes = nombre_completo.split()
    # Intento buscar combinaciones más amplias para nombres compuestos
    for i in range(1, len(partes)-1):
        for j in range(i+1, len(partes)):
            nombre = ' '.join(partes[:i])
            apellido_paterno = ' '.join(partes[i:j])
            apellido_materno = ' '.join(partes[j:])
            persona = db['datos_generales'].find_one({
                'nombre': {'$regex': f'^{nombre}$', '$options': 'i'},
                'apellido_paterno': {'$regex': f'^{apellido_paterno}$', '$options': 'i'},
                'apellido_materno': {'$regex': f'^{apellido_materno}$', '$options': 'i'}
            })
            if persona:
                return persona['RFC']
    return None

def procesar_datos_checador(fecha, nombre_completo, rfc):
    resultados = []
    formato_fecha_hora = '%Y-%m-%dT%H:%M:%S.%fZ'

    for hec in fecha.get('HEC', []):
        # Extraer 'hora_entrada' y verificar si ya es un objeto datetime o una cadena
        hora_entrada = hec.get('hora_entrada')
        if isinstance(hora_entrada, datetime):
            # Si es un objeto datetime, simplemente formatea la hora
            hora_entrada_dt = hora_entrada.isoformat()
        elif isinstance(hora_entrada, str) and hora_entrada:
            # Si es una cadena y no está vacía, intenta convertirla a datetime
            try:
                hora_entrada_dt = datetime.strptime(hora_entrada, formato_fecha_hora).isoformat()
            except ValueError:
                hora_entrada_dt = 'Hora inválida'
        else:
            # Si es una cadena vacía o no es ninguno de los tipos anteriores
            hora_entrada_dt = 'Hora no proporcionada'

        resultados.append({
            'RFC': rfc,
            'Fecha': fecha['fecha_dia'],
            'HS': hora_entrada_dt,
            'Estatus': hec.get('estatus_checador', 'Desconocido'),
            'Nombre': nombre_completo,
            'Tipo': 'Entrada'
        })

    for hsc in fecha.get('HSC', []):
        hora_salida = hsc.get('hora_salida')
        if isinstance(hora_salida, datetime):
            hora_salida_dt = hora_salida.isoformat()
        elif isinstance(hora_salida, str) and hora_salida:
            try:
                hora_salida_dt = datetime.strptime(hora_salida, formato_fecha_hora).isoformat()
            except ValueError:
                hora_salida_dt = 'Hora inválida'
        else:
            hora_salida_dt = 'Hora no proporcionada'

        resultados.append({
            'RFC': rfc,
            'Fecha': fecha['fecha_dia'],
            'HS': hora_salida_dt,
            'Estatus': hsc.get('estatus_checador', 'Desconocido'),
            'Nombre': nombre_completo,
            'Tipo': 'Salida'
        })

    return resultados



@app.route('/ventana_modificar')
def ventana_modificar():
    nombre_completo = request.args.get('Nombre')
    fecha = request.args.get('Fecha')
    tipo = request.args.get('Tipo')
    estatus = request.args.get('Estatus')

    rfc = obtener_rfc_por_nombre_completo(nombre_completo)
    if not rfc:
        return jsonify({'error': 'No se pudo encontrar un RFC para el nombre proporcionado'}), 404

    documento = db.catalogos_horario.find_one({"RFC": rfc, "Fechas.fecha_dia": fecha})
    horas_entrada = []
    horas_salida = []
    if documento:
        fecha_obj = next((item for item in documento['Fechas'] if item['fecha_dia'] == fecha), None)
        if fecha_obj:
            horas_entrada = [entry['hora_entrada'] for entry in fecha_obj.get('HEC', [])]
            horas_salida = [entry['hora_salida'] for entry in fecha_obj.get('HSC', [])]

    datos = {
        'Fecha': fecha,
        'Nombre': nombre_completo,
        'RFC': rfc,
        'HorasEntrada': horas_entrada,
        'HorasSalida': horas_salida,
        'Tipo': tipo,
        'Estatus': estatus
    }

    return render_template('ventana_modificar.html', datos=datos)






def convertir_hora_a_iso(hora_str):
    try:
        # Crear un objeto datetime combinando la fecha base con la hora
        fecha_base = datetime(1970, 1, 1)  # Año, Mes, Día
        hora_obj = datetime.strptime(hora_str, '%H:%M').time()
        datetime_completo = datetime.combine(fecha_base, hora_obj)

        # Convertir a UTC y luego a formato ISO 8601
        datetime_utc = pytz.utc.localize(datetime_completo)
        return datetime_utc.isoformat()
    except ValueError as e:
        print(f"Error al convertir la hora: {str(e)}")
        return None

@app.route('/actualizar_registro', methods=['POST'])
def actualizar_registro():
    rfc = request.form.get('rfc')
    fecha = request.form.get('fecha')
    hora_seleccionada = request.form.get('hora')  # Hora seleccionada desde el frontend
    estatus_nuevo = request.form.get('estatus')
    tipo = request.form.get('tipo')

    if not all([rfc, fecha, hora_seleccionada, estatus_nuevo, tipo]):
        return jsonify({"error": "Todos los campos son requeridos."}), 400

    documento = db['catalogos_horario'].find_one({"RFC": rfc, "Fechas.fecha_dia": fecha})
    if not documento:
        return jsonify({"mensaje": "No se encontró el documento correspondiente al RFC y fecha proporcionados."}), 404

    # Encontrar el objeto de fecha específico y actualizar solo el registro de hora seleccionado
    fecha_actualizada = False
    for fecha_obj in documento['Fechas']:
        if fecha_obj['fecha_dia'] == fecha:
            registros = fecha_obj['HEC'] if tipo == 'Entrada' else fecha_obj['HSC']
            for registro in registros:
                hora_actual = registro['hora_entrada'] if tipo == 'Entrada' else registro['hora_salida']
                # Compara la hora seleccionada con la hora del registro para encontrar la coincidencia
                if hora_actual == hora_seleccionada:
                    registro['estatus_checador'] = estatus_nuevo
                    fecha_actualizada = True
                    break

    if fecha_actualizada:
        resultado = db['catalogos_horario'].replace_one({"_id": documento['_id']}, documento)
        if resultado.modified_count > 0:
            return redirect(url_for('frameChecador'))
        else:
            return jsonify({"mensaje": "No se pudo actualizar el registro."}), 400
    else:
        return jsonify({"mensaje": "No se encontró la hora especificada en los registros."}), 404

    
    






@app.route('/frameReporteEmpleado', methods=['GET', 'POST'])
def frameReporteEmpleado():
    if request.method == 'POST':
        rfc_buscado = request.form.get('rfc')
        nombre_completo_buscado = request.form.get('nombre').strip()
        resultados = []

        if rfc_buscado:
            resultados = list(db['departamentos_area'].find({'RFC': rfc_buscado}))
        elif nombre_completo_buscado:
            partes = nombre_completo_buscado.split()
            # Generar posibles combinaciones de nombre, apellido paterno y materno
            for i in range(1, len(partes) - 1):
                for j in range(i + 1, len(partes)):
                    nombre = ' '.join(partes[:i])
                    apellido_paterno = ' '.join(partes[i:j])
                    apellido_materno = ' '.join(partes[j:])
                    persona = db['departamentos_area'].find_one({
                        'nombre': {'$regex': f'^{nombre}$', '$options': 'i'},
                        'apellido_paterno': {'$regex': f'^{apellido_paterno}$', '$options': 'i'},
                        'apellido_materno': {'$regex': f'^{apellido_materno}$', '$options': 'i'}
                    })
                    if persona:
                        resultados.append(persona)

        if not resultados:
            return render_template('reporte_empleado.html', resultados=[], error="No se encontraron resultados.")

        print("Resultados encontrados:", resultados)  # Depuración
        return render_template('reporte_empleado.html', resultados=resultados)
    else:
        return render_template('reporte_empleado.html')







#Esto es para mostrar en la tabla todos los empleados 
@app.route('/todosLosEmpleados')
def todosLosEmpleados():
    empleados = db['departamentos_area'].find()  # Asegúrate de que esta es la colección correcta
    lista_empleados = []

    for empleado in empleados:
        # Formatear el nombre completo del empleado
        nombre_completo = f"{empleado.get('nombre', '')} {empleado.get('apellido_paterno', '')} {empleado.get('apellido_materno', '')}".strip()
        # Asegurarse de que '_id' sea serializable
        empleado['_id'] = str(empleado['_id'])
        # Incluir el nombre completo en los datos del empleado
        empleado['nombre_completo'] = nombre_completo
        # Agregar el empleado modificado a la lista
        lista_empleados.append(empleado)

    return jsonify(lista_empleados)






#Frame para el Reporte Asistencia le modifique a esta parte 31/03/2024
@app.route('/frameReporteAsistencia', methods=['GET', 'POST'])
def frameReporteAsistencia():
    resultados = []
    if request.method == 'POST':
        rfc = request.form.get('rfc')
        nombre_completo = request.form.get('nombre').strip()
        todos = request.form.get('todos', type=bool, default=False)
        fecha_inicio = request.form.get('fecha_inicio')
        fecha_fin = request.form.get('fecha_fin') or fecha_inicio

        fecha_inicio_dt = fecha_fin_dt = None
        if fecha_inicio and fecha_fin:
            try:
                fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
                fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            except ValueError:
                fecha_inicio_dt = fecha_fin_dt = None

        query = {}
        if todos and fecha_inicio_dt and fecha_fin_dt:
            query = {}
        elif rfc:
            query = {"RFC": rfc}
        elif nombre_completo:
            query = obtener_query_por_nombre(nombre_completo)
            if not query:
                return render_template('reporte_asistencia.html', resultados=[], error="No se encontró el empleado con el nombre proporcionado.")

        if not query and not (fecha_inicio_dt and fecha_fin_dt):
            return render_template('reporte_asistencia.html', resultados=[], error="Especifica un rango de fechas o añade más filtros.")

        registros = db['catalogos_horario'].find(query)
        for registro in registros:
            justificaciones = list(db['justificaciones'].find({"RFC": registro.get('RFC', '')}))
            just_dict = parse_justificaciones(justificaciones)

            for fecha in registro.get('Fechas', []):
                fecha_dia_str = fecha.get('fecha_dia', "")
                if fecha_dia_str:
                    fecha_dia_dt = datetime.strptime(fecha_dia_str[:10], '%Y-%m-%d')
                    if fecha_inicio_dt and fecha_fin_dt:
                        if not (fecha_inicio_dt <= fecha_dia_dt <= fecha_fin_dt):
                            continue
                    procesar_datos(fecha, registro, resultados, just_dict, fecha_inicio_dt, fecha_fin_dt)

    return render_template('reporte_asistencia.html', resultados=resultados)

def obtener_query_por_nombre(nombre_completo):
    partes = nombre_completo.split()
    for i in range(1, len(partes)):
        for j in range(i + 1, len(partes) + 1):
            nombre = ' '.join(partes[:i])
            apellido_paterno = ' '.join(partes[i:j])
            apellido_materno = ' '.join(partes[j:])
            persona = db['datos_generales'].find_one({
                'nombre': {'$regex': f'^{nombre}$', '$options': 'i'},
                'apellido_paterno': {'$regex': f'^{apellido_paterno}$', '$options': 'i'},
                'apellido_materno': {'$regex': f'^{apellido_materno}$', '$options': 'i'}
            })
            if persona:
                return {'RFC': persona['RFC']}
    return None

def parse_justificaciones(justificaciones):
    just_dict = {}
    for j in justificaciones:
        fecha_inicial = parse_date(j['fecha_inicial'])
        just_dict[fecha_inicial] = j['incidencia']
        if 'fecha_final' in j:
            fecha_final = parse_date(j['fecha_final'])
            just_dict[fecha_final] = j['incidencia']
    return just_dict

def procesar_datos(fecha, registro, resultados, just_dict, fecha_inicio_dt, fecha_fin_dt):
    rfc = registro.get('RFC', '')
    persona = db['datos_generales'].find_one({'RFC': rfc})
    if persona:
        nombre_completo = " ".join([persona.get('nombre', 'Desconocido'), 
                                    persona.get('apellido_paterno', ''), 
                                    persona.get('apellido_materno', '')])
    else:
        nombre_completo = "Desconocido"

    fecha_dia_obj = parse_date(fecha['fecha_dia'])
    incidencia = just_dict.get(fecha_dia_obj, 'No Incidencias')

    horas_entrada = [format_datetime_or_string(hec['hora_entrada']) for hec in fecha.get('HEC', [])]
    horas_salida = [format_datetime_or_string(hsc['hora_salida']) for hsc in fecha.get('HSC', [])]

    estatus_checadores = [hec.get('estatus_checador', '---') for hec in fecha.get('HEC', [])] + \
                         [hsc.get('estatus_checador', '---') for hsc in fecha.get('HSC', [])]

    horas_totales_trabajadas = sum([(datetime.strptime(hs, '%H:%M') - datetime.strptime(he, '%H:%M')).seconds / 3600
                                    for he, hs in zip(horas_entrada, horas_salida) if datetime.strptime(hs, '%H:%M') > datetime.strptime(he, '%H:%M')])

    resultados.append({
        'Fecha': fecha_dia_obj.strftime('%Y-%m-%d') if isinstance(fecha_dia_obj, date) else fecha_dia_obj,
        'RFC': rfc,
        'Nombre': nombre_completo,
        'HE': ', '.join(horas_entrada),
        'HS': ', '.join(horas_salida),
        'HT': round(horas_totales_trabajadas, 2),
        'Incidencia': incidencia,
        'Estatus': ', '.join(estatus_checadores)
    })

def format_datetime_or_string(value):
    try:
        if isinstance(value, datetime):
            return value.strftime('%H:%M')
        # Intente convertir la cadena a datetime si no está vacía
        elif isinstance(value, str) and value:
            return datetime.strptime(value, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%H:%M')
        return value
    except ValueError:
        return value  # Retorna la cadena original si la conversión falla

def parse_date(date_str):
    try:
        return datetime.strptime(date_str[:10], '%Y-%m-%d').date()
    except ValueError:
        return date_str 



@app.route('/exportar-a-excel', methods=['POST'])
def exportar_a_excel():
    datos_json = request.get_json()
    df = DataFrame(datos_json.get('datos'), columns=['FECHA', 'RFC', 'NOMBRE', 'HORA ENTRADA', 'HORA SALIDA', 'HORAS TOTALES', 'INCIDENCIAS', 'ESTATUS'])

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, startcol=1)  # Comienza desde la columna B
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Configura el ancho de las columnas
        for col in range(2, 10):
            worksheet.column_dimensions[get_column_letter(col)].width = 20

        # Combinar celdas para títulos y aplicar estilos
        worksheet.merge_cells('B1:I2')
        worksheet.merge_cells('B3:I4')
        worksheet.merge_cells('B5:I6')

        # Configura los títulos
        titles = ["TECNOLÓGICO NACIONAL DE MÉXICO CAMPUS TUXTEPEC", "RECURSOS HUMANOS", "REGISTRO DE ASISTENCIA"]
        for i, title in enumerate(titles):
            cell = worksheet.cell(row=2*i+1, column=2)
            cell.value = title
            cell.font = Font(size=14, bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Configura los bordes para los títulos
        for row in range(1, 7):
            for col in range(2, 10):
                worksheet.cell(row=row, column=col).border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        # Configura las cabeceras de la tabla y define los bordes
        for col, header in enumerate(df.columns, start=2):
            cell = worksheet.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))
            cell.alignment = Alignment(horizontal="center")

        # Agrega los datos de la tabla y define los bordes
        for r_idx, data in enumerate(df.values, start=8):
            for c_idx, value in enumerate(data, start=2):
                cell = worksheet.cell(row=r_idx, column=c_idx)
                cell.value = value
                cell.border = Border(top=Side(style='thin'), bottom=Side(style='thin'), left=Side(style='thin'), right=Side(style='thin'))

        # Aplica el filtro a la fila de las cabeceras
        worksheet.auto_filter.ref = 'B7:I7'

    output.seek(0)
    return send_file(output, as_attachment=True, download_name="Reporte_Asistencia.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route('/exportar-a-excelem', methods=['POST'])
def exportar_a_excelem():
    datos_recibidos = request.get_json()
    search_terms_rfc = [d['RFC'] for d in datos_recibidos['datos'] if 'RFC' in d]

    # Campos específicos para cada colección
    campos_generales = ['RFC', 'nombre', 'apellido_paterno', 'apellido_materno', 'fecha_de_nacimiento', 
                        'edad', 'sexo', 'estado_civil', 'CURP', 'telefono', 'calle', 'numero_de_casa', 
                        'colonia', 'ciudad', 'codigo_postal']
    campos_contratacion = ['RFC', 'departamento_o_area', 'puesto']

    # Consulta a MongoDB utilizando los RFCs y campos específicos
    documentos_generales = list(db['datos_generales'].find({'RFC': {'$in': search_terms_rfc}}, {campo: 1 for campo in campos_generales}))
    documentos_contratacion = list(db['datos_contratacion'].find({'RFC': {'$in': search_terms_rfc}}, {campo: 1 for campo in campos_contratacion}))

    # Creación de DataFrames a partir de los documentos obtenidos
    df_generales = DataFrame(documentos_generales)
    df_contratacion = DataFrame(documentos_contratacion)

    # Eliminar duplicados considerando múltiples campos
    if 'S/RFC' in search_terms_rfc:
        df_generales = df_generales.drop_duplicates(subset=['nombre', 'apellido_paterno', 'apellido_materno'])
        df_contratacion = df_contratacion.drop_duplicates(subset=['RFC'])

    # Fusión de DataFrames basados en 'RFC'
    df_final = pd.merge(df_generales, df_contratacion, on='RFC', how='left')

    # Eliminar columnas no deseadas que podrían haber aparecido después de la fusión
    df_final.drop(columns=[col for col in df_final.columns if col.endswith('_x') or col.endswith('_y')], inplace=True)

    # Cambiar nombres de cabeceras
    nombres_cabeceras = ["RFC", "Nombre", "Apellido Paterno", "Apellido Materno", "Fecha de Nacimiento", "Edad", 
                         "Sexo", "Estado Civil", "CURP", "Telefono", "Calle", "Numero de Casa", "Colonia",
                         "Ciudad", "Codigo Postal", "Departamento/Area", "Puesto"]
    df_final.columns = nombres_cabeceras

    # Preparación del archivo Excel
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_final.to_excel(writer, index=False, startrow=6)
        workbook = writer.book
        worksheet = writer.sheets['Sheet1']

        # Aplicación de estilos al archivo Excel
        for col_num, col_name in enumerate(nombres_cabeceras, 1):
            worksheet.column_dimensions[get_column_letter(col_num)].width = 30

        ultima_columna = get_column_letter(len(nombres_cabeceras))
        worksheet.merge_cells(f'A1:{ultima_columna}2')
        worksheet['A1'] = 'TECNOLÓGICO NACIONAL DE MÉXICO CAMPUS TUXTEPEC'
        worksheet.merge_cells(f'A3:{ultima_columna}4')
        worksheet['A3'] = 'REPORTE DE EMPLEADOS'
        worksheet.merge_cells(f'A5:{ultima_columna}6')
        worksheet['A5'] = 'RECURSOS HUMANOS'

        # Añadir bordes a las filas combinadas
        combined_rows = [(1, 2), (3, 4), (5, 6)]
        for start_row, end_row in combined_rows:
            for row in worksheet.iter_rows(min_row=start_row, max_row=end_row, max_col=len(nombres_cabeceras)):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for row in worksheet.iter_rows(min_row=7, max_row=worksheet.max_row, max_col=len(nombres_cabeceras)):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    output.seek(0)
    return send_file(output, as_attachment=True, download_name="Reporte_Datos_Generales.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


#ESTAS RUTAS SON PARA EL REPORTE DE HORARIO, EMPIEZAN RUTAS DE REPORTE HORARIO

def convertir_hora(hora):
    """Convierte una cadena de fecha y hora ISO o un objeto datetime a un formato de hora legible HH:MM."""
    if isinstance(hora, datetime):
        return hora.strftime('%H:%M')
    elif isinstance(hora, str):
        try:
            if hora.endswith('Z'):
                hora = hora[:-1] + '+00:00'
            hora_dt = datetime.fromisoformat(hora)
            return hora_dt.strftime('%H:%M')
        except ValueError:
            return 'Formato inválido'
    else:
        return 'No registrada'

def obtener_query_por_nombre(nombre_completo):
    partes = nombre_completo.split()
    for i in range(1, len(partes)):
        for j in range(i + 1, len(partes) + 1):
            nombre = ' '.join(partes[:i])
            apellido_paterno = ' '.join(partes[i:j])
            apellido_materno = ' '.join(partes[j:])
            persona = db['datos_generales'].find_one({
                'nombre': {'$regex': f'^{nombre}$', '$options': 'i'},
                'apellido_paterno': {'$regex': f'^{apellido_paterno}$', '$options': 'i'},
                'apellido_materno': {'$regex': f'^{apellido_materno}$', '$options': 'i'}
            })
            if persona:
                return {'RFC': persona['RFC']}
    return None

@app.route('/frameReporteHorario', methods=['GET', 'POST'])
def frameReporteHorario():
    resultados = []
    error = None

    if request.method == 'POST':
        rfc = request.form.get('rfc')
        nombre_completo = request.form.get('nombre').strip()

        if not rfc and not nombre_completo:
            error = "Ingresa un dato para realizar la búsqueda."
        else:
            query = {}
            if rfc:
                query = {'RFC': rfc}
            elif nombre_completo:
                query = obtener_query_por_nombre(nombre_completo)

            if not query:
                error = "No se encontró el empleado."

            if query:
                empleados = db['catalogos_horario'].find(query)
                horario_abierto = True
                for empleado in empleados:
                    if 'Horarios' in empleado:
                        for horario in empleado['Horarios']:
                            if 'DIA' in horario:
                                horario_abierto = False
                                for dia, info in horario['DIA'].items():
                                    entradas = [convertir_hora(hora) for hora in info.get('Hora_entrada', [])]
                                    salidas = [convertir_hora(hora) for hora in info.get('Hora_salida', [])]
                                    resultados.append({
                                        'Dia': dia,
                                        'Entrada': ', '.join(entradas),
                                        'Salida': ', '.join(salidas)
                                    })
                if horario_abierto:
                    error = "ESTE ES UN HORARIO ABIERTO"

    return render_template('reporte_horario.html', resultados=resultados, error=error)


#TERMINAN RUTAS DE REPORTE HORARIO








#Frame para el boton de Cancelar del Reporte Checador
@app.route('/ventana_checador')
def ventana_checador():
    return render_template('checador.html')

@app.route('/checador')
def checador():
    return render_template('checador.html')

#Frame con el que se despliega la tabla de los reportes
@app.route('/frameArchivos')
def frameArchivos():
  return render_template('archivos.html') 






@app.route('/generar_pdf', methods=['POST'])
def generar_pdf():
   # Consulta a la base de datos
    datos_generales = list(db['datos_generales'].find())
    datos_contratacion = list(db['datos_contratacion'].find())

    # Renderiza el template HTML con los datos consultados
    html = render_template('reporte_empleado_pdf.html', datos_generales=datos_generales, datos_contratacion=datos_contratacion)

    # Configura pdfkit y genera el PDF
    config = pdfkit.configuration(wkhtmltopdf=r'C:\\Program Files\\wkhtmltopdf\\bin\\wkhtmltopdf.exe')
    pdf = pdfkit.from_string(html, False, configuration=config)

    # Crea y devuelve la respuesta con el PDF
    response = make_response(pdf)
    response.headers['Content-Type'] = 'application/pdf'
    response.headers['Content-Disposition'] = 'attachment; filename=reporte_empleados.pdf'

    return response

@app.route('/datos-empresa')
def datos_empresa():
    # Puedes pasar variables adicionales a render_template si es necesario
    return render_template('datos_empresa.html')

@app.route('/usuarios', methods=['GET', 'POST'])
def usuarios():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        apellido_paterno = request.form.get('apellido_paterno')
        apellido_materno = request.form.get('apellido_materno')
        usuario_id = request.form.get('usuario')  # Asegúrate de que esto coincida con "name" en el HTML
        contrasena = request.form.get('contraseña')  # Asegúrate de que esto coincida con "name" en el HTML
        institucion = request.form.get('institucion')

        nuevo_usuario = {
            "nombre": nombre,
            "apellido_paterno": apellido_paterno,
            "apellido_materno": apellido_materno,
            "ID": usuario_id,
            "contraseña": contrasena,
            "institucion": institucion
        }

        db['seguridadU'].insert_one(nuevo_usuario)

        return redirect(url_for('usuarios'))
    return render_template('usuarios.html')

#ESTAS SON LAS RUTAS PARA EL BOTON DE CHECADOR, EMPIEZAN RUTAS

def formatear_hora(hora):
    if isinstance(hora, datetime):
        return hora.strftime('%H:%M:%S')
    elif isinstance(hora, str):
        try:
            if hora.endswith('Z'):
                hora = hora[:-1] + '+00:00'
            fecha = datetime.fromisoformat(hora)
            return fecha.strftime('%H:%M:%S')
        except ValueError:
            return 'Formato inválido'
    else:
        return 'No registrada'

@app.route('/botonchecador')
def botonchecador():
    # Asegúrate de que estás tomando solo el último registro actualizado
    registro = db['catalogos_horario'].find({}).sort('ultima_actualizacion', -1)
    resultados = []
    estatus_permitidos = ["RETARDO", "NORMAL", "NOTA MALA"]  # Lista de estatus permitidos

    for reg in registro:
        rfc = reg['RFC']
        persona = db['datos_generales'].find_one({'RFC': rfc})
        if persona:
            nombre_completo = f"{persona['nombre']} {persona['apellido_paterno']} {persona['apellido_materno']}"
            for fecha in reg.get("Fechas", []):
                for hec in fecha.get("HEC", []):
                    if hec.get('estatus_checador') in estatus_permitidos:
                        resultados.append({
                            'fecha_dia': fecha['fecha_dia'],
                            'nombre': nombre_completo,
                            'hora': formatear_hora(hec['hora_entrada']),
                            'tipo': 'Entrada',
                            'estatus': hec['estatus_checador']
                        })
                for hsc in fecha.get("HSC", []):
                    if hsc.get('estatus_checador') in estatus_permitidos:
                        resultados.append({
                            'fecha_dia': fecha['fecha_dia'],
                            'nombre': nombre_completo,
                            'hora': formatear_hora(hsc['hora_salida']),
                            'tipo': 'Salida',
                            'estatus': hsc['estatus_checador']
                        })

    return render_template('botonchecador.html', resultados=resultados)

@app.route('/actualizaciones-recientes')
def actualizaciones_recientes():
    # Define el tiempo mínimo para considerar una actualización reciente
    tiempo_minimo = datetime.now(pytz.utc) - timedelta(seconds=20)
    
    # Encuentra registros que han sido actualizados recientemente y cuyo estatus ha cambiado
    registros = db['catalogos_horario'].find({
        'ultima_actualizacion': {'$gte': tiempo_minimo},
        'estatus_checador': {'$in': ['RETARDO', 'NORMAL', 'NOTA MALA']}
    }).sort('ultima_actualizacion', -1)
    
    resultados = []
    for reg in registros:
        rfc = reg['RFC']
        persona = db['datos_generales'].find_one({'RFC': rfc})
        if persona:
            nombre_completo = f"{persona['nombre']} {persona['apellido_paterno']} {persona['apellido_materno']}"
            for fecha in reg.get('Fechas', []):
                for hec in fecha.get('HEC', []):
                    resultados.append({
                        'fecha_dia': fecha['fecha_dia'],
                        'nombre': nombre_completo,
                        'hora': formatear_hora(hec['hora_entrada']),
                        'tipo': 'Entrada',
                        'estatus': hec['estatus_checador']
                    })
                for hsc in fecha.get('HSC', []):
                    resultados.append({
                        'fecha_dia': fecha['fecha_dia'],
                        'nombre': nombre_completo,
                        'hora': formatear_hora(hsc['hora_salida']),
                        'tipo': 'Salida',
                        'estatus': hsc['estatus_checador']
                    })

    return jsonify(resultados)



#TERMINAN RUTAS DE BOTON CHECADOR

@app.route('/frameinstitucion')
def institucion():
    # Puedes pasar variables adicionales a render_template si es necesario
    return render_template('institucion.html')

@app.route('/autocompleterep', methods=['GET'])
def autocompleterep():
    tipo = request.args.get('tipo', 'nombre')  # Recoge el tipo de búsqueda: 'nombre' o 'rfc'
    search = request.args.get('term', '')

    if tipo == 'nombre':
        # Búsqueda por nombre, incluyendo todos los componentes del nombre
        cursor = db['datos_generales'].find({
            '$or': [
                {'nombre': {'$regex': f'^{search}', '$options': 'i'}},
                {'apellido_paterno': {'$regex': f'^{search}', '$options': 'i'}},
                {'apellido_materno': {'$regex': f'^{search}', '$options': 'i'}}
            ]
        }, {'nombre': 1, 'apellido_paterno': 1, 'apellido_materno': 1, '_id': 0}).limit(10)

        result_list = [{'label': f"{x['nombre']} {x['apellido_paterno']} {x['apellido_materno']}",
                        'value': f"{x['nombre']} {x['apellido_paterno']} {x['apellido_materno']}"} for x in cursor]

    elif tipo == 'rfc':
        # Búsqueda por RFC
        cursor = db['datos_generales'].find({'RFC': {'$regex': f'^{search}', '$options': 'i'}}, {'RFC': 1}).limit(10)
        result_list = [{'label': x['RFC'], 'value': x['RFC']} for x in cursor]

    return jsonify(result_list)




#TERMINAN RUTAS DE TOÑO---------------------------------------------------------------------NO MODIFICAR-------------------------------------------------------------------------------------------


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        usuario = request.form['usuario']
        contraseña = request.form['contraseña']
        # Aquí deberías verificar las credenciales con tu base de datos.
        usuario_db = db['seguridadU'].find_one({"ID": usuario})
        if usuario_db and usuario_db['contraseña'] == contraseña:
            # Si el usuario y la contraseña son correctos, redirigir al index.html
            return redirect(url_for('principal'))
        else:
            # Si son incorrectos, mostrar un mensaje de error en la página de login
            error = "Usuario o contraseña incorrectos"
            return render_template('login.html', error=error)
    # Si el método es GET, simplemente renderiza la página de inicio de sesión
    return render_template('login.html')



# INICIA RUTAS ALBERTO LOGIN

@app.route('/login1')
def login1():
    # Captura el mensaje de error si existe
    error = request.args.get('error', None)
    return render_template('login.html', error=error)

@app.route('/verificar_rfc', methods=['POST'])
def verificar_rfc():
    rfc_ingresado = request.form['RFC']
    if rfc_valido(rfc_ingresado):
        return redirect(url_for('pagina_jefesin'))
    else:
        return redirect(url_for('login1', error='RFC no válido'))

def rfc_valido(rfc):
    """ Verifica si el RFC existe en la colección 'login'. """
    collection = db['login']  # Asegúrate de que 'login' es el nombre correcto de tu colección
    rfc_encontrado = collection.find_one({'RFC': rfc})  # La búsqueda en MongoDB es sensible a mayúsculas y minúsculas
    return rfc_encontrado is not None  # Retorna True si se encontró el RFC, de lo contrario False

@app.route('/pagina_jefesin')
def pagina_jefesin():
    justificaciones = list(db.Justificaciones.find({}))
    justificaciones_enriquecidas = []
    #FALTA CONSEGUIR LOS DEPARTAMENTOS O AREAS PARA HACER UNA BUSQUEDA DE LOS QUE TENGAN EL MISMO AREA O DEPARTAMENTO
    for justificacion in justificaciones:
        rfc = justificacion.get('RFC')  # Asumiendo que cada justificación tiene un campo RFC
        empleado = db.datos_generales.find_one({'RFC': rfc})

        if empleado:
            # Enriquece la información de justificación con detalles del empleado
            justificacion['dg_apellido_paterno'] = empleado.get('apellido_paterno', 'No disponible')
            justificacion['dg_apellido_materno'] = empleado.get('apellido_materno', 'No disponible')
        else:
            # Si no encuentra el empleado, establece los campos como no disponibles
            justificacion['dg_apellido_paterno'] = 'No disponible'
            justificacion['dg_apellido_materno'] = 'No disponible'

        justificaciones_enriquecidas.append(justificacion)

    return render_template('jefeincidencias.html', justificaciones=justificaciones_enriquecidas)


# FINALZIA RUTAS ALBERTO

import logging
from logging.handlers import RotatingFileHandler

# Configuración del logging
logging.basicConfig(level=logging.INFO)  # Puedes cambiar a DEBUG, ERROR, etc., según lo que necesites
handler = RotatingFileHandler('index.log', maxBytes=1000000, backupCount=2)
handler.setLevel(logging.INFO)  # O cualquier otro nivel según necesites
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger = logging.getLogger(__name__)
logger.addHandler(handler)




#ALIOTH RUTAS :)
#inicia rutas AÑLIOTHHHHHHHHHH
#Rutas de la aplicacion

@app.route('/eliminarHorario/<nombre_reporte>', methods=['POST'])
def eliminar_horario(nombre_reporte):
    # MongoDB usa el operador $pull para eliminar un elemento del array que coincida con ciertos criterios
    resultado = db.catalogos_horario.update_many(
        {},
        {'$pull': {'Horarios': {'Nombre_horario_reporte': nombre_reporte}}}
    )
    
    if resultado.modified_count > 0:
        # Reporte eliminado con éxito
        return jsonify({'mensaje': 'Reporte eliminado exitosamente'}), 200
    else:
        # Reporte no encontrado o no eliminado por alguna razón
        return jsonify({'mensaje': 'No se encontró el reporte o no se pudo eliminar'}), 404



from datetime import datetime

@app.route('/verHorario/<nombre_reporte>')
def ver_horario(nombre_reporte):
    reportes = db.catalogos_horario.find({'Horarios.Nombre_horario_reporte': nombre_reporte})
    reporte_encontrado = None
    rfc_empleado = None

    # Buscar el reporte específico y el RFC del empleado
    for reporte in reportes:
        for horario in reporte['Horarios']:
            if horario['Nombre_horario_reporte'] == nombre_reporte:
                reporte_encontrado = horario
                rfc_empleado = reporte['RFC']
                break
        if reporte_encontrado:
            break

    if not reporte_encontrado or not rfc_empleado:
        return "Reporte o empleado no encontrado", 404

    empleado = db.datos_generales.find_one({'RFC': rfc_empleado})
    contrato = db.datos_contratacion.find_one({'RFC': rfc_empleado}) or {}
    nombre_completo = f"{empleado.get('nombre')} {empleado.get('apellido_paterno')} {empleado.get('apellido_materno')}"
            
    if not empleado:
        return "Información de empleado no encontrada", 404

    nombre_dia_a_numero = {'LUNES': 1, 'MARTES': 2, 'MIERCOLES': 3, 'JUEVES': 4, 'VIERNES': 5, 'SABADO': 6, 'DOMINGO': 7}

    # Procesar cada día en el horario encontrado
    horario_procesado = {}
    for dia_nombre, horas in reporte_encontrado.get('DIA', {}).items():
        dia_numero = nombre_dia_a_numero.get(dia_nombre.upper())
        if dia_numero:
            horario_procesado[dia_numero] = {
                'Hora_entrada': [datetime.strptime(hora, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%H:%M') for hora in horas.get('Hora_entrada', [])],
                'Hora_salida': [datetime.strptime(hora, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%H:%M') for hora in horas.get('Hora_salida', [])],
            }
        print(horario_procesado)
   


    datos_horario = {
        'estatus': 'Activo' if reporte_encontrado.get('estatus', '') == 'Activo' else 'Inactivo',
        'horario_actual': reporte_encontrado.get('Nombre_horario_reporte', 'No disponible'),
        'horario': horario_procesado,
        'nombre': nombre_completo,
        'puesto': contrato.get('puesto', 'Puesto no disponible'),
        'fecha_reporte': reporte_encontrado.get('Fecha_reporte', 'Fecha no disponible')
    }

    datos_contratacion = {
        'fecha_de_ingreso': contrato.get('fecha_de_ingreso', 'Fecha de ingreso no disponible'),
        'departamento_o_area': contrato.get('departamento_o_area', 'Departamento no disponible'),
        'puesto': contrato.get('puesto', 'Puesto no disponible')
    }

    # Renderiza la plantilla HTML con los datos del reporte
    return render_template('horarioshowreporte.html', datos_horario=datos_horario, datos_contratacion=datos_contratacion, datos_generales=empleado)

@app.route('/horariotabla/<rfc>')
def horariotabla(rfc):
    empleado = db.datos_generales.find_one({'RFC': rfc})
    if not empleado:
        return "Empleado no encontrado", 404

    contrato = db.datos_contratacion.find_one({'RFC': rfc}) or {}
    fecha_ingreso = contrato.get('fecha_de_ingreso', 'Fecha de ingreso no disponible')
    departamento = contrato.get('departamento_o_area', 'Departamento no disponible')
    puesto = contrato.get('puesto', 'Puesto no disponible')

    horario = db.catalogos_horario.find_one({'RFC': rfc})
    if horario and 'Horarios' in horario and horario['Horarios']:
        ultimo_reporte = max(horario['Horarios'], key=lambda x: x['Fecha_reporte'])
        fecha_carga = ultimo_reporte['Fecha_reporte']
        Nombre_horario_reporte = ultimo_reporte['Nombre_horario_reporte']
        
        nombre_dia_a_numero = {'LUNES': 1, 'MARTES': 2, 'MIERCOLES': 3, 'JUEVES': 4, 'VIERNES': 5, 'SABADO': 6, 'DOMINGO': 7}

        horario_procesado = {}
        for dia_nombre, horas in ultimo_reporte.get('DIA', {}).items():
            dia_numero = nombre_dia_a_numero.get(dia_nombre.upper())
            if dia_numero:
                horario_procesado[dia_numero] = {
                    'Hora_entrada': [datetime.strptime(hora, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%H:%M') for hora in horas.get('Hora_entrada', [])],
                    'Hora_salida': [datetime.strptime(hora, '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%H:%M') for hora in horas.get('Hora_salida', [])],
                }
            print(horario_procesado)

        datos_horario = {
            'nombre': empleado.get('nombre', ''),
            'apellidos': f"{empleado.get('apellido_paterno', '')} {empleado.get('apellido_materno', '')}",
            'fecha_carga': fecha_carga,
            'puesto': puesto,
            'estatus': 'ACTIVO' if empleado.get('estatus', '') == 'Activo' else 'INACTIVO',
            'horario': horario_procesado,
            'horario_actual': Nombre_horario_reporte
        }
    else:
        datos_horario = {
            'nombre': empleado.get('nombre', ''),
            'apellidos': f"{empleado.get('apellido_paterno', '')} {empleado.get('apellido_materno', '')}",
            'fecha_carga': 'N/H',
            'puesto': puesto,
            'estatus': 'ACTIVO' if empleado.get('estatus', '') == 'Activo' else 'INACTIVO',
            'horario': {},
            'ultimo_reporte': '...'
        }

    return render_template('horariocarga.html', datos_horario=datos_horario, fecha_ingreso=fecha_ingreso, departamento=departamento, datos_contratacion=contrato, datos_generales=empleado)

###################################
from datetime import datetime, timezone
@app.route('/guardarHorario', methods=['PUT'])
def guardar_horario():
    horario_data = request.get_json()
    rfc = horario_data.get('RFC')
    puesto = horario_data.get('puesto')
    empleado = horario_data.get('empleado')
    horarios = horario_data.get('horarios')
    estatus = horario_data.get('estatus')
    tipo_horario = horario_data.get('tipo_horario')
    fechas_dia = horario_data.get('fecha_dias')  # Obtener las fechas del JSON

    # Procesar y convertir las horas de entrada y salida
    for dia, dia_info in horarios.items():
        dia_info['Hora_entrada'] = [datetime.strptime('1970-02-01T' + hora + ':00.000Z', '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%Y-%m-%dT%H:%M:%S.%f')[:23] + 'Z' for hora in dia_info.get('Hora_entrada', [])]
        dia_info['Hora_salida'] = [datetime.strptime('1970-02-01T' + hora + ':00.000Z', '%Y-%m-%dT%H:%M:%S.%fZ').strftime('%Y-%m-%dT%H:%M:%S.%f')[:23] + 'Z' for hora in dia_info.get('Hora_salida', [])]
    # Actualizar o insertar el horario en la base de datos
    horario_existente = db.catalogos_horario.find_one({'RFC': rfc})
    if horario_existente:
        nuevo_horario = {
            'Nombre_horario_reporte': f'{rfc}_{len(horario_existente["Horarios"]) + 1}',
            'Fecha_reporte': datetime.now(timezone.utc),
            'puesto': puesto,
            'estatus': estatus,
            'empleado': empleado
        }
        
        # Si el tipo de horario es "Abierto", no se agrega información de horarios específicos para cada día
        if tipo_horario == 'Abierto':
            nuevo_horario['DIA'] = {}  # Se establece como un diccionario vacío
        else:
            nuevo_horario['DIA'] = horarios
        
        db.catalogos_horario.update_one({'RFC': rfc}, {'$push': {'Horarios': nuevo_horario}, '$set': {'tipo_horario': tipo_horario}})
        
        # Agregar nuevas fechas y actualizar las existentes
        if fechas_dia:
            for fecha_nueva in fechas_dia:
                if tipo_horario == 'Cerrado':
                    fecha_nueva['HEC'] = [{'hora_entrada': datetime.strptime(hec['hora_entrada'], '%Y-%m-%dT%H:%M:%S.%fZ').replace(tzinfo=timezone.utc), 'estatus_checador': 'FALTA'} for hec in fecha_nueva.get('HEC', [])]
                    fecha_nueva['HSC'] = [{'hora_salida': datetime.strptime(hsc['hora_salida'], '%Y-%m-%dT%H:%M:%S.%fZ').replace(tzinfo=timezone.utc), 'estatus_checador': 'FALTA'} for hsc in fecha_nueva.get('HSC', [])]
                elif tipo_horario == 'Abierto':
                    fecha_nueva['HEC'] = [{'hora_entrada': None, 'estatus_checador': 'ABIERTO'} for _ in fecha_nueva.get('HEC', [])]
                    fecha_nueva['HSC'] = [{'hora_salida': None, 'estatus_checador': 'ABIERTO'} for _ in fecha_nueva.get('HSC', [])]
                db.catalogos_horario.update_one({'RFC': rfc, 'Fechas.fecha_dia': {'$ne': fecha_nueva['fecha_dia']}}, {'$addToSet': {'Fechas': fecha_nueva}})
                db.catalogos_horario.update_one({'RFC': rfc, 'Fechas.fecha_dia': fecha_nueva['fecha_dia']}, {'$set': {'Fechas.$': fecha_nueva}})

    else:
        nuevo_horario = {
            'RFC': rfc,
            'Horarios': [{
                'Nombre_horario_reporte': f'{rfc}_1',
                'Fecha_reporte': datetime.now(timezone.utc),
                'puesto': puesto,
                'estatus': estatus,
                'empleado': empleado
            }],
            'tipo_horario': tipo_horario
        }
        
        # Si el tipo de horario es "Abierto", no se agrega información de horarios específicos para cada día
        if tipo_horario != 'Abierto':
            nuevo_horario['Horarios'][0]['DIA'] = horarios
        
        if fechas_dia:
            for fecha_nueva in fechas_dia:
                if tipo_horario == 'Cerrado':
                    fecha_nueva['HEC'] = [{'hora_entrada': datetime.strptime(hec['hora_entrada'], '%Y-%m-%dT%H:%M:%S.%fZ').replace(tzinfo=timezone.utc), 'estatus_checador': 'FALTA'} for hec in fecha_nueva.get('HEC', [])]
                    fecha_nueva['HSC'] = [{'hora_salida': datetime.strptime(hsc['hora_salida'], '%Y-%m-%dT%H:%M:%S.%fZ').replace(tzinfo=timezone.utc), 'estatus_checador': 'FALTA'} for hsc in fecha_nueva.get('HSC', [])]
                elif tipo_horario == 'Abierto':
                    fecha_nueva['HEC'] = [{'hora_entrada': None, 'estatus_checador': 'ABIERTO'} for _ in fecha_nueva.get('HEC', [])]
                    fecha_nueva['HSC'] = [{'hora_salida': None, 'estatus_checador': 'ABIERTO'} for _ in fecha_nueva.get('HSC', [])]
        nuevo_horario['Fechas'] = fechas_dia
        db.catalogos_horario.insert_one(nuevo_horario)
        
    return jsonify({'mensaje': 'Horario guardado exitosamente'}), 200






#Frame con el que se despliega reportes de horario, principal en catalogo-horario- despliega esto...
#Estas rutas se llaman en la parte de index.html o tambien en otros template dependiendo el caso... en este caso 
#solo devuelve la plantilla de horarios.html pero tambien puede hacer una solicitud get o post o etc...


db = dbase.dbConnection()
@app.route('/framehorario')
def framehorario():
    reportes_horario = db.catalogos_horario.find()

    # Se crea una lista que contendrá todos los horarios con la información complementaria
    horarios_completos = []
    
    for reporte in reportes_horario:
        rfc = reporte.get('RFC')
        
        datos_generales = db.datos_generales.find_one({'RFC': rfc})
        datos_contratacion = db.datos_contratacion.find_one({'RFC': rfc})
        
        # Verifica si el empleado existe en datos_generales y datos_contratacion
        if datos_generales and datos_contratacion:
            # Aquí se extrae el 'estatus' de datos_generales y 'puesto' de datos_contratacion
            estatus = datos_generales.get('estatus','Desconocido')
            puesto = datos_contratacion.get('puesto', 'Desconocido')
            nombre_completo = f"{datos_generales.get('nombre')} {datos_generales.get('apellido_paterno')} {datos_generales.get('apellido_materno')}"
            
            # Invertir el orden de los horarios
            horarios_invertidos = list(reversed(reporte.get('Horarios', [])))
            
            # Crea un nuevo diccionario con toda la información
            horario_completo = {
                'RFC': rfc,
                'Horarios': horarios_invertidos,
                'estatus': estatus,
                'puesto': puesto,
                'nombrecompleto': nombre_completo
            }
            horarios_completos.append(horario_completo)

    # Enviar la lista de horarios completos a la plantilla
    return render_template('horarios.html', reportes_horario=horarios_completos)



@app.route('/framehorarioEmpleados')
def framehorarioEmpleados():
    # Obtener los datos de la colección 'datos_generales'
    datos_generales = list(db.datos_generales.find())

    # Obtener los datos de la colección 'datos_contratacion'
    datos_contratacion = list(db.datos_contratacion.find())

    # Convertir datos_generales en un diccionario para un acceso más rápido
    generales_dict = {item['RFC']: item for item in datos_generales}

    # Combinar los datos
    empleados = []
    for contrato in datos_contratacion:
        rfc = contrato['RFC']
        empleado = generales_dict.get(rfc)
        if empleado:
            empleado.update({
                'Puesto': contrato.get('puesto', ''),  # Corregir a 'puesto'
                'Fecha_Contratación': contrato.get('fecha_de_ingreso', ''),  # Corregir a 'fecha_de_ingreso'
                'estatus': empleado.get('estatus', '')  # Usar el estatus de datos_generales
            })
            empleados.append(empleado)

    return render_template('horarioempleados.html', empleados=empleados)


import logging

#FINALIZA ROUTES ALIOTH

@app.route('/frameRed')
def frameRed():
  return render_template('red.html') 


def ping(host):
    parametro = '-n' if platform.system().lower() == 'windows' else '-c'
    comando = ['ping', parametro, '2', host]
    result = subprocess.run(comando, capture_output=True, text=True)
    print(result.stdout)  # Imprime la salida para depuración
    return result.stdout


@app.route('/ping/<host>')
def ping_host(host):
    output = ping(host)
    print("Output del ping:", output)  # Agrega esto para depurar
    if 'tiempo=' in output or 'time=' in output:  # Agrega 'tiempo=' por si acaso
        return jsonify(status="success", result=output)
    else:
        return jsonify(status="failure", result=output)



#PoUp Se Agrego Correctamente
@app.route('/empleadoExitoso')
def empleadoExitoso():
  return render_template('empleadoExitoso.html')


#Frame con Form Agregar Empleado
@app.route('/frameEmpleados')
def frameEmpleados():
  empleados = db['datos_generales']
  empleadosReceived = empleados.find()
  return render_template('empleados.html')

@app.route('/empleadoEliminadoExitoso')
def empleadoEliminadoExitoso():
  return render_template('empleadoEliminadoExitoso.html')

@app.route('/empleadoEditadoExitoso')
def empleadoEditadoExitoso():
  return render_template('empleadoEditadoExitoso.html')

@app.route('/empleadoEliminadoError')
def empleadoEliminadoError():
  return render_template('empleadoEliminadoError.html')

@app.route('/empleadoEditadoError')
def empleadoEditadoError():
  return render_template('empleadoEditadoError.html')


#Metodo Agregar empleados
@app.route('/addEmpleados', methods={'POST'})
def addEmpleado():
  empleados_dg = db['datos_generales']
  empleados_dc = db['datos_contratacion']

  rfc = request.form['rfc']
  nombre = request.form['nombre']
  apellido_p = request.form['apellido_p']
  apellido_m = request.form['apellido_m']
  fecha_de_nac = request.form['fecha_de_nac']
  edad = request.form['edad']
  sexo = request.form['sexo']
  edo_civil = request.form['edo_civil']
  curp = request.form['curp']
  telefono = request.form['telefono']
  calle = request.form['calle']
  #num_casa = request.form['num_casa']
  colonia = request.form['colonia']
  ciudad = request.form['ciudad']
  estado = request.form['estado']
  cod_pos = request.form['cod_pos']
  num_creden = request.form['num_creden']
  tipo_horario = request.form['tipo_horario']
  estatus = request.form['estatus']
  departamento = request.form.get('departamentos')
  puesto = request.form.get('puestos')
  id_horario = request.form['id_horario']
  fecha_contratacion = request.form['fecha_contratacion']
  salario = request.form['salario']
  no_ss = request.form['no_ss']

#Agregar de nuevo a la lista el numero de casa
  if rfc and apellido_p and apellido_m:
    empleado = Empleado(rfc, nombre, apellido_p,apellido_m,fecha_de_nac, edad, sexo, edo_civil, curp, telefono, calle, colonia, ciudad, estado, cod_pos, num_creden, tipo_horario,estatus)
    empleados_dg.insert_one(empleado.insertDatosGenerales())
    empleados_dc.insert_one({'RFC': rfc, 'nombre': nombre, 'apellido_paterno': apellido_p, 'apellido_materno': apellido_m, 'departamento_o_area': departamento, 'puesto': puesto, 'id_horario': id_horario, 'fecha_de_ingreso': fecha_contratacion, 'seguro_social': no_ss, 'salario': salario})
    response = jsonify({
      '_id': rfc,
      'nombre' : nombre,
      'apellido_p': apellido_p,
      'apellido_m': apellido_m,
      'fecha_de_nacimiento': fecha_de_nac,
      'edad': edad,
      'sexo': sexo,
      'estado_civil': edo_civil,
      'CURP': curp,
      'telefono': telefono,
      'calle': calle,
      #'numero_de_casa': num_casa,
      'colonia': colonia,
      'ciudad': ciudad,
      'estado': estado,
      'codigo_postal': cod_pos,
      'numero_de_credencial': num_creden,
      'tipo_de_horario': tipo_horario,
      'estatus': estatus
      
    })
    return redirect(url_for('empleadoExitoso'))
  else:
    return notFound()

#Metodo Delete
@app.route('/deleteEmpleados', methods={'POST'})
def delete():
  empleado_rfc = request.form['rfc']

  empleados_dg = db['datos_generales']
  empleados_dc = db['datos_contratacion']
  empleados_dg.delete_one({'RFC': empleado_rfc})
  empleados_dc.delete_one({'RFC': empleado_rfc})

  if empleados_dg.count_documents({'RFC':empleado_rfc}) == 0 and empleados_dc.count_documents({'RFC':empleado_rfc}) == 0: 
    return redirect(url_for('empleadoEliminadoExitoso'))
  else:
    return redirect(url_for('empleadoEliminadoError'))

#Metodo Put
@app.route('/editEmpleados', methods=['POST'])
def edit():
  empleados_dg = db['datos_generales']
  empleados_dc = db['datos_contratacion']

  empleado_rfc = request.form['rfc']
  nombre = request.form['nombre']
  ape_pat = request.form['apellido_p']
  ape_mat = request.form['apellido_m']
  fecha_nac = request.form['fecha_de_nac']
  edad = request.form['edad']
  sexo = request.form['sexo']
  edo_civil = request.form['edo_civil']
  curp = request.form['curp']
  tel = request.form['telefono']
  calle = request.form['calle']
  #num_casa = request.form['num_casa']
  col = request.form['colonia']
  ciudad = request.form['ciudad']
  edo = request.form['estado']
  cp = request.form['cod_pos']
  num_cred = request.form['num_creden']
  tipo_hor = request.form['tipo_horario']
  estatus = request.form['estatus']
  departamento = request.form.get('departamentos')
  puesto = request.form.get('puestos')
  id_horario = request.form['id_horario']
  fecha_contratacion = request.form['fecha_contratacion']
  no_ss = request.form['no_ss']

  fecha_nac_split = fecha_nac.split('-')
  fecha_nac_dia = int(fecha_nac_split[0])
  fecha_nac_mes = int(fecha_nac_split[1])
  fecha_nac_ano= int(fecha_nac_split[2])
  fecha_nac_format = datetime(fecha_nac_ano, fecha_nac_mes, fecha_nac_dia)


  fecha_cont_split = fecha_contratacion.split('-')
  fecha_cont_dia = int(fecha_cont_split[0])
  fecha_cont_mes = int(fecha_cont_split[1])
  fecha_cont_ano= int(fecha_cont_split[2])
  fecha_cont_format = datetime(fecha_cont_ano, fecha_cont_mes, fecha_cont_dia)

  empleado_dg = empleados_dg.find_one({ 'RFC': {'$eq': empleado_rfc }})
  
  #Lista que se crea, para despues usarla como llaves de un diccionario
  campos = ['rfc', 'nombre', 'ape_pat', 'ape_mat', 'fech_nac', 'edad', 'sexo','edo_civ', 'curp', 'tel','calle', 'col', 'ciudad', 'edo', 'cp', 'num_cred', 'tipo_hor', 'estatus']
  #Valores actuales del Empleado con el RFC que tecleo el usuario, y se guardaran al recorrer el documento de la BD
  valoresActuales = []
  for i in empleado_dg:
    valoresActuales.append(empleado_dg[i])
  #Se crea un diccionario, para despues acceder a él de manera mas inmediata a los valores mediante las llaves creadas
  empleadoActual = dict(zip(campos, valoresActuales))
  #print(empleadoActual['nombre']) #Sólo para validar que si esta modificando el empleado deseado

  if empleado_rfc == empleadoActual['rfc']:
    if nombre != '' and ape_pat != '' and ape_mat != '' and fecha_nac != '' and edad != '' and sexo != '' and edo_civil != '' and curp != '' and tel != '' and calle != '' and col != '' and ciudad != '' and edo != '' and cp != '' and num_cred != '' and tipo_hor != '' and estatus != '':
      empleados_dg.find_one_and_update({'RFC': empleado_rfc},{'$set': {'RFC': empleado_rfc, 'nombre': nombre, 'apellido_paterno': ape_pat, 'apellido_materno': ape_mat, 'fecha_de_nacimiento': fecha_nac_format,'edad': edad, 'sexo': sexo, 'estado_civil': edo_civil, 'CURP': curp,'telefono': tel, 'calle': calle, 'colonia': col, 'ciudad': ciudad, 'estado': edo, 'codigo_postal': cp, 'numero_de_credencial': num_cred, 'tipo_de_horario': tipo_hor, 'estatus': estatus }})
      empleados_dc.find_one_and_update({'RFC': empleado_rfc}, {'$set': {'RFC': empleado_rfc, 'nombre':nombre, 'apellido_paterno': ape_pat, 'apellido_materno': ape_mat, 'departamento_o_area': departamento, 'puesto': puesto, 'secuencia_de_horario': id_horario, 'fecha_de_ingreso': fecha_cont_format, 'seguro_social': no_ss}})
      response = jsonify({'message' : 'Empleado ' + empleado_rfc + ' actualizado correctamente'})  
      return redirect(url_for('empleadoEditadoExitoso'))
    else:
       return redirect(url_for('empleadoEditadoError'))
  else:
    return redirect(url_for('empleadoEditadoRfc'))


#API para autocompletado
@app.route('/autocomplete', methods=['GET', 'POST'])
def autocomplete():

  empleados = db['datos_generales']
  cursor = empleados.find({}, {'_id':0})
  #resultado = []
  #for index, empleado in enumerate(cursor):
    #nombre = empleado["nombre"]
    #apellido_pat = empleado["apellido_paterno"]
    #apellido_mat = empleado["apellido_materno"]
  return jsonify(list(cursor))

#Endpoint Busqueda de empleados 
@app.route('/searchEmpleados', methods=['GET', 'POST'])
def search():

  empleados = db['datos_generales']
  empleados_con = db['datos_contratacion']
    
  tipo = request.form.get("busqueda")
  clave = request.form['buscador']


  carpeta = 'C:/Users/siste/Documents/PROYECTO_CHECADOR/sistemaChecador_flask/static/db/imagenes'

  #Obtener una lista de todos los archivos en la carpeta
  archivos = os.listdir(carpeta)

  #Filtrar solo los archivos que son imágenes (puedes agregar más extensiones según tus necesidades)
  imagenes = [archivo for archivo in archivos if archivo.endswith(('.jpg'))]
   
  imagen = ''
  img_rfc = ''
  empleadoReceived = ''
  contratacionReceived = ''
  if tipo == 'rfc' and tipo != 'nombre':
    empleadoReceived = empleados.find({ 'RFC': clave})
    contratacionReceived = empleados_con.find({'RFC': clave})
    
    clave_img = clave+".jpg"
    #Recorrer todas las imágenes que estan en la carpeta
    for imagen in imagenes:
      #print(imagen)
      if clave_img == imagen:
        print("Imagen con "+clave_img + " encontrada")
        img_rfc = imagen

  elif tipo == 'nombre' and tipo != 'rfc':
    i = 0
    espacios = 0

    while i < len(clave):
      if clave[i] == " ":
         espacios = espacios + 1
      i = i + 1
    #Verificamos si solo tiene un nombre, comparando los espacios del nombre completo
    #Si el nombre completo ingresado en la busqueda tiene un nombre, este tiene solo 2 espacios
    if espacios == 2:
      nombre_completo = clave.split()
      nombre = nombre_completo[0]
      apellido_pat = nombre_completo[1]
      apellido_mat = nombre_completo[2]
      empleadoReceived = empleados.find({'$and': [{'nombre': nombre}, {'apellido_paterno': apellido_pat}, {'apellido_materno': apellido_mat}]})
      contratacionReceived = empleados_con.find({'$and': [{'nombre': nombre}, {'apellido_paterno': apellido_pat}, {'apellido_materno': apellido_mat}]})
        
      rfc = ''
      empleado_seleccionado = empleados.find({'$and': [{'nombre': nombre}, {'apellido_paterno': apellido_pat}, {'apellido_materno': apellido_mat}]})
      for empleado in empleado_seleccionado:
        rfc = empleado['RFC']
      
      clave_img = rfc+".jpg"
      #Recorrer todas las imágenes que estan en la carpeta
      for imagen in imagenes:
        print(imagen)
        if clave_img == imagen:
          print("Imagen con "+clave_img + " encontrada")
          img_rfc = imagen
        else:
          print("Imagen no encontrada con el RFC: " +rfc)


    #Pero si el nombre completo tiene 3 espacios, significa que tiene 2 nombres y hay que guardarlos en una sola variable
    elif espacios == 3:
      nombre_completo = clave.split()
      nombre = nombre_completo[0] + ' ' + nombre_completo[1]
      apellido_pat = nombre_completo[2]
      apellido_mat = nombre_completo[3]
      empleadoReceived = empleados.find({'$and': [{'nombre': nombre}, {'apellido_paterno': apellido_pat}, {'apellido_materno': apellido_mat}]})
      contratacionReceived = empleados_con.find({'$and': [{'nombre': nombre}, {'apellido_paterno': apellido_pat}, {'apellido_materno': apellido_mat}]})
      
      rfc = ''
      empleado_seleccionado = empleados.find({'$and': [{'nombre': nombre}, {'apellido_paterno': apellido_pat}, {'apellido_materno': apellido_mat}]})
      for empleado in empleado_seleccionado:
        rfc = empleado['RFC']
      
      clave_img = rfc+".jpg"
      #Recorrer todas las imágenes que estan en la carpeta
      for imagen in imagenes:
        print(imagen)
        if clave_img == imagen:
          print("Imagen con "+clave_img + " encontrada")
          img_rfc = imagen
        else:
          print("Imagen no encontrada con el RFC: " +rfc)
  else:
    return print('No valido')

  return render_template('empleados.html', busqueda = [empleadoReceived, contratacionReceived, img_rfc]) 
 
######################################DEPARTAMENTO##########################################

# Crear el índice de texto en la colección 'department'
db.department.create_index([("name", "text")])

# Rutas de la aplicación

@app.route('/frameDepartment')
def frameDepartment():
    department = db['department']
    departmentReceived = department.find()
    return render_template('department.html', departments=departmentReceived)


# Nueva ruta para mostrar solo los departamentos
@app.route('/onlydepartment')
def onlyDepartment():
    department = db['department']
    departmentReceived = department.find()
    return render_template('onlydepartment.html', departments=departmentReceived)

# Método Post
@app.route('/department', methods=['POST'])
def addDepartment():
    departments = db['department']
    cid = request.form['cid']
    name = request.form['name']
    
    # Verificar si se proporcionaron tanto el CID como el nombre
    if not cid or not name:
        return "<script>alert('Por favor, complete tanto la clave como el nombre del departamento.'); window.location.href = '/frameDepartment';</script>", 400
    
    # Verificar si el CID tiene exactamente 5 caracteres y contiene una combinación de letras, números y caracteres especiales
    if len(cid) != 5 or not any(char.isalpha() for char in cid) or not any(char.isdigit() for char in cid) or not all(char.isalnum() or char in ['_', '-'] for char in cid):
        return "<script>alert('La clave debe tener exactamente 5 caracteres alfanuméricos. También puedes usar _ y -'); window.location.href = '/frameDepartment';</script>", 400
    
    # Verificar si el CID ya existe en la base de datos
    existing_department_cid = departments.find_one({'cid': cid})
    if existing_department_cid:
        # Si el CID ya existe, devuelve un mensaje de error
        return "<script>alert('La clave ya está en uso'); window.location.href = '/frameDepartment';</script>", 400
    
    # Verificar si el nombre del departamento ya existe en la base de datos
    existing_department_name = departments.find_one({'name': name})
    if existing_department_name:
        # Si el nombre del departamento ya existe, devuelve un mensaje de error
        return "<script>alert('El nombre del departamento ya está en uso'); window.location.href = '/frameDepartment';</script>", 400
    
    # Si el CID no existe y es válido, y si el nombre del departamento no existe, proceder con la inserción del departamento
    if cid and name:
        department = Department(cid, name)
        departments.insert_one(department.toDBCollection())
        return redirect(url_for('frameDepartment'))
    else:
        return notFound()

# Método Delete

@app.route('/department/<string:department_cid>', methods=['DELETE'])
def delete_department(department_cid):
    department = db['department']
    department.delete_one({'cid': department_cid})
    return jsonify({'message': 'Departamento eliminado correctamente'}), 200


# Método Put
@app.route('/editDepartment/<string:department_cid>', methods=['POST'])
def editDepartment(department_cid):
    department = db['department']
    cid = request.form['cid']
    name = request.form['name']
    
    # Verificar si el CID tiene exactamente 5 caracteres y contiene una combinación de letras, números y caracteres especiales,
    # y al menos una letra
    if len(cid) != 5 or not any(char.isalpha() for char in cid) or not all(char.isalnum() or char in ['_', '-'] for char in cid):
        return "<script>alert('La clave debe tener exactamente 5 caracteres alfanuméricos con al menos una letra. También puedes usar _ y -'); window.location.href = '/frameDepartment';</script>", 400
    
    # Verificar si el CID ya existe en la base de datos
    existing_department_cid = department.find_one({'cid': cid})
    if existing_department_cid and cid != department_cid:
        # Si el CID ya existe y es diferente al CID original, devuelve un mensaje de error
        return "<script>alert('La clave ya está en uso'); window.location.href = '/frameDepartment';</script>", 400
    
    # Verificar si el nombre ya existe en la base de datos, excepto si pertenece al departamento actual
    existing_department_name = department.find_one({'name': name, 'cid': {'$ne': department_cid}})
    if existing_department_name:
        # Si el nombre ya existe para otro departamento, devuelve un mensaje de error
        return "<script>alert('El nombre ya está en uso'); window.location.href = '/frameDepartment';</script>", 400
    
    # Si el CID y el nombre son válidos, proceder con la actualización del departamento
    if cid and name:
        department.update_one({'cid': department_cid}, {'$set': {'name': name, 'cid': cid}})
        return redirect(url_for('frameDepartment'))
    else:
        return notFound()

# Ruta para buscar departamentos
@app.route('/searchDepartment')
def search_departments():
    query = request.args.get('query')
    department_collection = db['department']
    # Realiza la búsqueda en la base de datos y devuelve los resultados en formato JSON
    if query:
        search_results = department_collection.find({'$text': {'$search': query}})
    else:
        search_results = department_collection.find()
    results = [{'cid': department['cid'], 'name': department['name']} for department in search_results]
    return jsonify(results)

#######################################################################PUESTOS######################################
# Ruta para mostrar los puestos de trabajo
@app.route('/job_positions')
def show_job_positions():
    job_positions = db['job_positions']
    positions = job_positions.find()
    return render_template('job_positions.html', job_positions=positions, position_id=positions)

@app.route('/onlyjobpositions')
def onlyJobpositions():
    job_positions = db['job_positions']
    job_positionsReceived = job_positions.find()
    return render_template('onlyjobpositions.html', job_positions=job_positionsReceived)

# Método POST para agregar un nuevo puesto de trabajo
@app.route('/job_positions', methods=['POST'])
def add_job_position():
    job_positions = db['job_positions']
    title = request.form['title']
    position_id = request.form['position_id']  # Captura el position_id del formulario
    
    # Verificar si se proporcionó tanto el título como el ID del puesto de trabajo
    if not title or not position_id:
        return "<script>alert('Por favor, complete tanto el título como el ID del puesto de trabajo.'); window.location.href = '/job_positions';</script>", 400
    
    # Verificar si el position_id tiene exactamente 5 caracteres y contiene una combinación de letras y números
    if len(position_id) != 5 or not any(char.isalpha() for char in position_id) or not any(char.isdigit() for char in position_id):
        return "<script>alert('El ID del puesto debe tener exactamente 5 caracteres alfanuméricos.'); window.location.href = '/job_positions';</script>", 400
    
    # Verificar si el position_id ya existe en la base de datos
    existing_position_id = job_positions.find_one({'position_id': position_id})
    if existing_position_id:
        # Si el position_id ya existe, devuelve un mensaje de error
        return "<script>alert('El ID del puesto ya está en uso.'); window.location.href = '/job_positions';</script>", 400
    
    # Si el título y el ID del puesto son válidos y únicos, proceder con la inserción del puesto
    if title and position_id:
        job_position = JobPosition(title, position_id)
        job_positions.insert_one(job_position.toDBCollection())
        return redirect(url_for('show_job_positions'))
    else:
        return notFound()


# Método DELETE para eliminar un puesto de trabajo
@app.route('/job_positions/<string:position_id>', methods=['DELETE'])
def delete_job_position(position_id):
    job_positions = db['job_positions']
    job_positions.delete_one({'position_id': position_id})
    return jsonify({'message': 'Puesto de trabajo eliminado correctamente'}), 200

# Método para editar un puesto de trabajo
@app.route('/edit_JobPosition/<string:position_id>', methods=['POST'])
def edit_JobPosition(position_id):
    job_positions = db['job_positions']
    new_position_id = request.form['position_id']
    title = request.form['title']

    # Verificar si el nuevo título ya existe en otro puesto de trabajo
    existing_title = job_positions.find_one({'title': title, 'position_id': {'$ne': position_id}})
    if existing_title:
        return "<script>alert('El título del puesto ya está en uso'); window.location.href = '/job_positions';</script>", 400

    if new_position_id and title:  # Cambiado position_id a new_position_id
        job_positions.update_one({'position_id': position_id}, {'$set': {'position_id': new_position_id, 'title': title}})
        return redirect(url_for('show_job_positions'))
    else:
        return notFound()

    
# Ruta para buscar puestos de trabajo
@app.route('/search_job_positions')
def search_job_positions():
    query = request.args.get('query')
    job_positions_collection = db['job_positions']
    
    # Realiza la búsqueda en la base de datos y devuelve los resultados en formato JSON
    if query:
        search_results = job_positions_collection.find({'$text': {'$search': query}})
    else:
        search_results = job_positions_collection.find()
    
    # Formatea los resultados en un formato JSON
    results = [{'title': job_position['title'], 'position_id': job_position['position_id']} for job_position in search_results]
    return jsonify(results)




##################################DIAS FESTIVOS############################################################
#Frame Dias Festivos
@app.route('/diasFestivos')
def diasFestivos():
  
  dias = db['dias_festivos']
  diasReceived = dias.find()

  return render_template('dias_festivos.html', resultado = diasReceived)

#Metodo Agregar Dias Festivos
@app.route('/addDiasFestivos', methods={'POST'})
def addDiasFestivos():

  diasFestivos = db['dias_festivos']
  horarios = db['catalogos_horario']

  tipo = request.form['tipo']
  fecha_inicio = request.form['fecha_inicio']
  fecha_fin = request.form['fecha_fin']
  descripcion = request.form['descripcion']
  busqueda_fecha = fecha_inicio + "T00:00:00.000Z"

  if tipo and fecha_inicio and descripcion and not fecha_fin:
    diasFestivos.insert_one({'tipo': tipo,'fecha_inicio':fecha_inicio,'fecha_fin':fecha_inicio,'descripcion': descripcion})
    #Actualiza el campus estatus checador de Falta a NO LABORABLE en el dia insertado
    horarios.update_many({"Fechas": {"$elemMatch": {"fecha_dia": busqueda_fecha}}}, {"$set": {"Fechas.$[fecha].HEC.$[].estatus_checador": "NO LABORABLE"}}, array_filters= [{ "fecha.fecha_dia": busqueda_fecha}])
  
  elif tipo and fecha_inicio and fecha_fin and descripcion:
    diasFestivos.insert_one({'tipo': tipo,'fecha_inicio':fecha_inicio,'fecha_fin':fecha_fin,'descripcion': descripcion})

    inicio_format = fecha_inicio.split('-')
    fin_format = fecha_fin.split('-')
    inicio = datetime(int(inicio_format[0]), int(inicio_format[1]), int(inicio_format[2]))
    fin = datetime(int(fin_format[0]), int(fin_format[1]), int(fin_format[2]))
    lista_fechas = [(inicio + timedelta(days=d)).strftime("%Y-%m-%d"+"T00:00:00.000Z")for d in range((fin - inicio).days + 1)]
    #print(lista_fechas)
     
    for fecha in lista_fechas: 
      horarios.update_one({"Fechas": {"$elemMatch": {"fecha_dia": fecha}}}, {"$set": {"Fechas.$[fecha].HEC.$[].estatus_checador": "NO LABORABLE"}}, array_filters = [{ "fecha.fecha_dia": fecha}])
  else:
    #Mostrar PopUp para que le muestre el error que no se pasaron todos los datos
    print('Inserte todos los datos')
  return redirect(url_for('diasFestivos'))


@app.route('/deleteDiasFestivos', methods={'GET', 'POST'})
def deleteDiasFestivos():

  diasFestivos = db['dias_festivos']
  horarios = db['catalogos_horario']

  tipo = request.args.get('tipo')
  fecha_inicio = request.args.get('fecha_inicio')
  fecha_fin = request.args.get('fecha_fin')
  descripcion = request.args.get('descripcion')

  if tipo and fecha_inicio and fecha_fin and descripcion:
    inicio_format = fecha_inicio.split('-')
    fin_format = fecha_fin.split('-')
    inicio = datetime(int(inicio_format[0]), int(inicio_format[1]), int(inicio_format[2]))
    fin = datetime(int(fin_format[0]), int(fin_format[1]), int(fin_format[2]))
    lista_fechas = [(inicio + timedelta(days=d)).strftime("%Y-%m-%d"+"T00:00:00.000Z")for d in range((fin - inicio).days + 1)]
    
    for fecha in lista_fechas:  
      horarios.update_one({"Fechas": {"$elemMatch": {"fecha_dia": fecha}}}, {"$set": {"Fechas.$[fecha].HEC.$[].estatus_checador": "FALTA"}}, array_filters = [{ "fecha.fecha_dia": fecha}])
    diasFestivos.delete_one({'descripcion': descripcion})
  else:
    #Mostrar PopUp para que le muestre el error que no se pasaron todos los datos
    print('Selecciona un dia a eliminar')

  return redirect(url_for('diasFestivos'))


@app.route('/ejecutarRegistro', methods={'GET', 'POST'})
def registro():
    exec(open("/registro_facial_huella/register_user.py").read())
    return redirect(url_for('frameEmpleados'))

@app.errorhandler(404)
def notFound(error=None):
  message ={
    'message': 'No encontrado',
    'status': '404 Not Found'
  }
  response= jsonify(message)
  response.status_code = 404
  return response


if __name__ == '__main__':
  app.run(debug=True, port=5000, host='0.0.0.0', threaded=True, use_reloader=True)